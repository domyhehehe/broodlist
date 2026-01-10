import argparse
import csv
import os
import sys


def load_rows(csv_path):
    with open(csv_path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
    by_pk = {}
    for row in rows:
        pk = (row.get("PrimaryKey") or "").strip()
        if not pk:
            continue
        by_pk[pk] = row
    return by_pk


def compute_max_depth(pk, by_pk):
    memo = {}
    visiting = set()

    def depth_for(node_pk):
        if node_pk in memo:
            return memo[node_pk]
        if node_pk in visiting:
            return 0
        visiting.add(node_pk)
        data = by_pk.get(node_pk)
        if data is None:
            memo[node_pk] = 0
            visiting.remove(node_pk)
            return 0
        sire = (data.get("Sire") or "").strip()
        dam = (data.get("Dam") or "").strip()
        depths = [0]
        if sire:
            depths.append(1 + depth_for(sire))
        if dam:
            depths.append(1 + depth_for(dam))
        max_depth = max(depths)
        memo[node_pk] = max_depth
        visiting.remove(node_pk)
        return max_depth

    return depth_for(pk)


def clamp_depth(pk, by_pk, requested_depth):
    max_depth = compute_max_depth(pk, by_pk)
    if requested_depth is None:
        return max_depth
    return max(0, min(max_depth, requested_depth))


def display_name(pk, data, note=""):
    if data is None:
        base = pk
        suffix = note or "not_found"
        return f"{base} ({suffix})"
    name = (data.get("Horse Name") or "").strip()
    year = (data.get("Year") or "").strip()
    if not name:
        base = pk
    else:
        base = name
    if year:
        base = f"{base} {year}"
    if base == pk:
        return base
    return f"{base} ({pk})"


def format_horse_name_year(pk, by_pk):
    data = by_pk.get(pk)
    if data is None:
        return pk
    name = (data.get("Horse Name") or "").strip() or pk
    year = (data.get("Year") or "").strip()
    return f"{name} {year}".strip()


def collect_inbreeding(pk, by_pk, max_depth):
    occurrences = {}

    def walk(node_pk, gen, path, visiting):
        if gen > max_depth:
            return
        if node_pk in visiting:
            return
        visiting.add(node_pk)
        new_path = path + [node_pk]
        occurrences.setdefault(node_pk, []).append({"gen": gen, "path": new_path})
        data = by_pk.get(node_pk)
        if data is None:
            visiting.remove(node_pk)
            return
        sire = (data.get("Sire") or "").strip()
        dam = (data.get("Dam") or "").strip()
        if gen < max_depth:
            if sire:
                walk(sire, gen + 1, new_path, visiting)
            if dam:
                walk(dam, gen + 1, new_path, visiting)
        visiting.remove(node_pk)

    data = by_pk.get(pk)
    if data:
        sire = (data.get("Sire") or "").strip()
        dam = (data.get("Dam") or "").strip()
        if sire:
            walk(sire, 1, [], set())
        if dam:
            walk(dam, 1, [], set())

    inbred = {}
    for ancestor, occs in occurrences.items():
        if len(occs) < 2:
            continue
        gens = [occ["gen"] for occ in occs]
        percentage = sum((0.5**g) for g in gens) * 100.0
        gens_sorted = sorted(gens, reverse=True)
        paths = [occ["path"] for occ in occs]
        inbred[ancestor] = {"gens": gens_sorted, "percentage": percentage, "paths": paths}

    subsumed = set()
    candidates = list(inbred.keys())
    for ancestor in candidates:
        for other in candidates:
            if ancestor == other:
                continue
            all_contained = True
            for path in inbred[ancestor]["paths"]:
                if other not in path:
                    all_contained = False
                    break
                if path.index(other) > path.index(ancestor):
                    all_contained = False
                    break
            if all_contained:
                subsumed.add(ancestor)
                break

    return inbred, subsumed


def write_excel(pk, by_pk, out_path, max_depth):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required. Install with: pip install -r requirements.txt"
        ) from exc

    total_rows = 2**max_depth if max_depth > 0 else 1
    excel_row_limit = 1_048_576
    if total_rows + 2 > excel_row_limit:
        raise SystemExit(
            f"Too many rows for Excel ({total_rows} > {excel_row_limit})."
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Pedigree"

    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_align = Alignment(horizontal="left", vertical="center")
    sire_fill = PatternFill(fill_type="solid", fgColor="FFDDEBF7")
    dam_fill = PatternFill(fill_type="solid", fgColor="FFFCE4EC")
    red_side = Side(style="thin", color="FFFF0000")
    red_border = Border(left=red_side, right=red_side, top=red_side, bottom=red_side)
    for col_idx in range(1, max_depth + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 26

    inbred, subsumed = collect_inbreeding(pk, by_pk, max_depth)
    inbred_pks = set(inbred.keys()) - subsumed

    title = format_horse_name_year(pk, by_pk)
    ws.merge_cells(
        start_row=1,
        end_row=1,
        start_column=1,
        end_column=max_depth if max_depth > 0 else 1,
    )
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.alignment = header_align

    if inbred_pks:
        summary_parts = []
        for ancestor_pk in sorted(
            inbred_pks, key=lambda pk: inbred[pk]["percentage"], reverse=True
        ):
            data = by_pk.get(ancestor_pk)
            name = (data.get("Horse Name") or "").strip() if data else ancestor_pk
            gens_text = " x ".join(str(g) for g in inbred[ancestor_pk]["gens"])
            summary_parts.append(
                f"{name} {inbred[ancestor_pk]['percentage']:.2f}% {gens_text}"
            )
        summary = " / ".join(summary_parts)
    else:
        summary = "No inbreeding detected within selected generations."

    ws.merge_cells(
        start_row=2,
        end_row=2,
        start_column=1,
        end_column=max_depth if max_depth > 0 else 1,
    )
    header_cell = ws.cell(row=2, column=1, value=summary)
    header_cell.alignment = header_align

    visiting = set()

    def place_node(node_pk, gen, row_start, row_end):
        data = by_pk.get(node_pk)
        if node_pk in visiting:
            text = display_name(node_pk, data, note="cycle")
        else:
            text = display_name(node_pk, data)

        col = gen + 1
        cell = ws.cell(row=row_start + 3, column=col, value=text)
        cell.alignment = align
        sex = (data.get("Sex") or "").strip().upper() if data else ""
        if sex == "M":
            cell.fill = sire_fill
        elif sex == "F":
            cell.fill = dam_fill
        if node_pk in inbred_pks:
            cell.border = red_border
        if row_end > row_start:
            ws.merge_cells(
                start_row=row_start + 3,
                end_row=row_end + 3,
                start_column=col,
                end_column=col,
            )

        if data is None or gen >= max_depth - 1:
            return

        sire = (data.get("Sire") or "").strip()
        dam = (data.get("Dam") or "").strip()
        if not sire and not dam:
            return

        mid = (row_start + row_end) // 2
        visiting.add(node_pk)
        if sire:
            place_node(sire, gen + 1, row_start, mid)
        if dam:
            place_node(dam, gen + 1, mid + 1, row_end)
        visiting.remove(node_pk)

    if max_depth > 0:
        data = by_pk.get(pk)
        sire = (data.get("Sire") or "").strip() if data else ""
        dam = (data.get("Dam") or "").strip() if data else ""
        if sire or dam:
            mid = (total_rows - 1) // 2
            if sire:
                place_node(sire, 0, 0, mid)
            if dam:
                place_node(dam, 0, mid + 1, total_rows - 1)

    wb.save(out_path)


def parse_args():
    parser = argparse.ArgumentParser(
        description="Create an Excel pedigree file from bloodline.csv"
    )
    parser.add_argument("pk", nargs="?", help="PrimaryKey")
    parser.add_argument(
        "--csv",
        default="bloodline.csv",
        help="Path to input CSV (default: bloodline.csv)",
    )
    parser.add_argument(
        "--gen",
        type=int,
        default=None,
        help="Max generations (default: prompt, fallback to 5)",
    )
    parser.add_argument(
        "--out",
        default=None,
        help="Output Excel path (default: <PK>.xlsx in current directory)",
    )
    return parser.parse_args()


def main():
    args = parse_args()
    pk = (args.pk or "").strip()
    if not pk:
        pk = input("Enter PrimaryKey: ").strip()
    if not pk:
        raise SystemExit("PrimaryKey is required.")

    csv_path = args.csv
    if not os.path.exists(csv_path):
        raise SystemExit(f"CSV not found: {csv_path}")

    by_pk = load_rows(csv_path)
    if pk not in by_pk:
        raise SystemExit(f"PrimaryKey not found in CSV: {pk}")

    gen = args.gen
    if gen is None:
        gen_input = input("Enter generations (default 5): ").strip()
        if gen_input:
            try:
                gen = int(gen_input)
            except ValueError as exc:
                raise SystemExit("Generations must be an integer.") from exc
        else:
            gen = 5
    if gen < 0:
        raise SystemExit("Generations must be >= 0.")

    max_depth = clamp_depth(pk, by_pk, gen)

    out_path = args.out or f"{pk}.xlsx"
    write_excel(pk, by_pk, out_path, max_depth)
    print(f"Wrote: {out_path}")


if __name__ == "__main__":
    main()
